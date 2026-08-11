from pathlib import Path
import pandas as pd
from typing import List, Dict, Any

from src.parsers.parser_factory import (
    SheetKind, 
    detect_sheet_family, 
    get_parser_for_sheet,
    infer_variable_semantics,
    should_include_in_market_dataset
)
from src.normalization.normalizer import merge_long_tables


def ingest_workbook(path: str, required_variables: set[str] = None) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """Ingest Excel workbook and return unified normalized DataFrame with report.

    Args:
        path: Path to workbook
        required_variables: Optional set of required variable names (e.g., {'Cours', 'Bid', 'Ask'})
    
    Returns:
        (unified_dataframe, ingestion_report)
    """
    pathp = Path(path)
    # Use openpyxl engine explicitly for better compatibility
    try:
        xls = pd.ExcelFile(pathp, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Failed to open Excel file: {e}. Ensure file is a valid .xlsx format.")
    
    long_tables: List[pd.DataFrame] = []
    
    report = {
        'total_sheets': len(xls.sheet_names),
        'sheets_processed': [],
        'sheets_included': [],
        'sheets_excluded': [],
        'warnings': []
    }

    for sheet in xls.sheet_names:
        sheet_report = {
            'name': sheet,
            'kind': None,
            'canonical_variable': None,
            'confidence': None,
            'included': False,
            'reason': None,
            'records': 0
        }
        
        try:
            # Load a small sample (first 20 rows) to detect sheet family and headers.
            # NOTE: using a small `nrows` keeps detection fast and low-memory even for very large workbooks.
            # The full sheet is parsed later by the selected parser (e.g. `parse_simple_sheet`) which
            # reads only the required parts for normalization. For UI previews, use `read_sheet_page()` below.
            df = pd.read_excel(pathp, sheet_name=sheet, nrows=20, header=None, engine='openpyxl')
            sheet_kind = detect_sheet_family(df)
            sheet_report['kind'] = sheet_kind.value
            
            # Infer semantic meaning
            canonical_variable, confidence = infer_variable_semantics(df, sheet)
            sheet_report['canonical_variable'] = canonical_variable
            sheet_report['confidence'] = confidence
            
            # Decide whether to include
            should_include, reason = should_include_in_market_dataset(
                sheet_kind, 
                canonical_variable, 
                confidence,
                required_variables
            )
            sheet_report['reason'] = reason
            
            if should_include:
                parser = get_parser_for_sheet(df)
                if parser is not None:
                    parsed = parser(pathp, sheet)
                    long_tables.append(parsed)
                    sheet_report['included'] = True
                    sheet_report['records'] = len(parsed)
                    report['sheets_included'].append(sheet_report)
                    print(f'✓ Included: {sheet} -> {canonical_variable} ({len(parsed)} records)')
                else:
                    sheet_report['reason'] = 'No parser available'
                    report['sheets_excluded'].append(sheet_report)
                    print(f'✗ Excluded: {sheet} -> No parser available')
            else:
                report['sheets_excluded'].append(sheet_report)
                print(f'⊗ Excluded: {sheet} -> {reason}')
                
        except Exception as e:
            sheet_report['reason'] = f'Error: {type(e).__name__}: {str(e)}'
            report['sheets_excluded'].append(sheet_report)
            report['warnings'].append(f'{sheet}: {e}')
            print(f'⚠ Warning: failed to process sheet {sheet}: {e}')
        
        report['sheets_processed'].append(sheet_report)

    unified = merge_long_tables(long_tables)
    
    report['unified_records'] = len(unified)
    report['unified_companies'] = unified['CODE_ISIN'].nunique() if not unified.empty else 0
    report['unified_sessions'] = unified['Date'].nunique() if not unified.empty else 0
    report['unified_variables'] = list(unified.columns.difference(['Date', 'CODE_ISIN', 'Company'])) if not unified.empty else []
    
    return unified, report


def get_sheet_row_count(path: str, sheet: str) -> int:
    """Return the number of rows in an Excel sheet using openpyxl without loading full data.

    This is a lightweight way to determine whether a sheet is "large" and to support pagination
    when previewing contents in the UI.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        # openpyxl is required as engine elsewhere; surface a clear error if missing
        raise

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return 0
    ws = wb[sheet]
    return ws.max_row


def read_sheet_page(path: str, sheet: str, page: int = 0, page_size: int = 10) -> tuple[pd.DataFrame, int]:
    """Read a single page (slice) of an Excel sheet without loading the entire sheet.

    Args:
        path: Path to workbook
        sheet: Sheet name
        page: Zero-based page index
        page_size: Number of rows per page

    Returns:
        (DataFrame of the requested slice, total_rows)

    Notes:
        - This function uses `skiprows` + `nrows` to limit memory footprint for previewing.
        - It is intended for UI previews (e.g. Streamlit) and not for full ingestion.
    """
    pathp = Path(path)
    total = get_sheet_row_count(pathp, sheet)
    # pandas.skiprows expects 0-based row indices to skip. We skip the first `start` rows.
    start = page * page_size
    if start >= total:
        return pd.DataFrame(), total

    # Read the requested window. header=None to preserve raw rows for preview.
    df = pd.read_excel(
        pathp,
        sheet_name=sheet,
        skiprows=start,
        nrows=page_size,
        header=None,
        engine='openpyxl'
    )
    return df, total


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('Usage: python -m src.ingestion <workbook.xlsx>')
        sys.exit(1)
    
    # Example with optional allowlist
    required_vars = {'Cours', 'Bid', 'Ask', 'Volume MC', 'Quantité MC'}
    df, report = ingest_workbook(sys.argv[1], required_variables=required_vars)
    
    print('\n=== INGESTION REPORT ===')
    print(f"Total sheets: {report['total_sheets']}")
    print(f"Included: {len(report['sheets_included'])}")
    print(f"Excluded: {len(report['sheets_excluded'])}")
    print(f"Unified records: {report['unified_records']}")
    print(f"Companies: {report['unified_companies']}")
    print(f"Sessions: {report['unified_sessions']}")
    print(f"Variables: {report['unified_variables']}")
    
    if report['warnings']:
        print(f"\nWarnings: {len(report['warnings'])}")
        for w in report['warnings']:
            print(f"  - {w}")
    
    out = Path('data/processed/unified_sample.parquet')
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out)
    print(f'\nWrote {out}')
